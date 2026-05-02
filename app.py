"""
Huliot CAD Quantity Extractor
Usage: python3 extract_qty.py <path_to_dxf_file> [--scale 1:100]

Extracts block INSERT quantities + pipe line counts from a DXF file.
Outputs:
  - Console summary
  - quantity_takeoff_<name>.xlsx
  - blocks_raw.csv
"""

import sys
import os
import re
import math
from collections import Counter, defaultdict
from pathlib import Path

try:
    import ezdxf
except ImportError:
    print("Installing ezdxf...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "ezdxf", "openpyxl",
                    "--break-system-packages", "-q"])
    import ezdxf

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


# ─────────────────────────────────────────────────────────────────────────────
# BLOCK NAME PARSER
# ─────────────────────────────────────────────────────────────────────────────

PRODUCT_PATTERNS = [
    (r"US_|ULTRA.?SILENT",       "Ultra Silent"),
    (r"PAP_|HELIROMA|PERT.AL",   "PERT-AL-PERT / Heliroma"),
    (r"HULIOT_|HTP_|HT.?PRO",    "HT Pro"),
]

TYPE_PATTERNS = [
    (r"ELBOW|ELB",               "Elbow"),
    (r"TEE|WYE|Y.BRANCH",        "Tee / Wye"),
    (r"REDUCER|REDUC|REDUCTI",   "Reducer"),
    (r"TRAP",                     "Trap"),
    (r"WC.?CON|WC|P.TRAP",       "WC Connector"),
    (r"ACCESS|CLEAN.?OUT|INSPEC", "Access Door"),
    (r"CAP|PLUG|END",             "End Cap"),
    (r"CLAMP|BRACKET|CLIP",       "Clamp"),
    (r"PIPE|SLEEVE",              "Pipe"),
    (r"COUPLING|JOINT|SOCKET",    "Coupling"),
    (r"MANIFOLD|HEADER",          "Manifold"),
    (r"FLOOR.?DRAIN|FD",          "Floor Drain"),
]

DN_SIZES = {"50", "56", "63", "75", "90", "110", "125", "160", "200", "250", "315"}


def _parse_sizes(name_up: str) -> list:
    """Extract DN sizes from block name, ignoring angle values (45°, 87°, 135°)."""
    nums = re.findall(r"\d+", name_up)
    dns  = [n for n in nums if n in DN_SIZES]
    # If 90 appears alongside other DN sizes, it is likely an angle (90°), not DN90
    if "90" in dns and len(dns) > 1:
        dns = [n for n in dns if n != "90"]
    return list(dict.fromkeys(dns))  # deduplicate, preserve order


def parse_block_name(name: str) -> dict:
    name_up = name.upper()

    # Product line
    product = "Unknown"
    for pattern, label in PRODUCT_PATTERNS:
        if re.search(pattern, name_up):
            product = label
            break

    # Type
    fitting_type = "Other"
    for pattern, label in TYPE_PATTERNS:
        if re.search(pattern, name_up):
            fitting_type = label
            break

    # Sizes
    sizes = _parse_sizes(name_up)
    size_str = "×".join(sizes) + " mm" if sizes else "—"

    return {
        "product":  product,
        "type":     fitting_type,
        "size":     size_str,
        "raw_name": name,
    }


# ─────────────────────────────────────────────────────────────────────────────
# LINE LENGTH
# ─────────────────────────────────────────────────────────────────────────────

def line_length(entity) -> float:
    """Return length of LINE or LWPOLYLINE entity."""
    try:
        if entity.dxftype() == "LINE":
            s = entity.dxf.start
            e = entity.dxf.end
            return math.sqrt((e[0]-s[0])**2 + (e[1]-s[1])**2)
        elif entity.dxftype() == "LWPOLYLINE":
            pts = list(entity.get_points())
            total = 0.0
            for i in range(len(pts) - 1):
                total += math.sqrt((pts[i+1][0]-pts[i][0])**2 + (pts[i+1][1]-pts[i][1])**2)
            return total
    except Exception:
        return 0.0
    return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# MAIN EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract(dxf_path: str, drawing_scale: float = 100.0) -> dict:
    """
    Extract quantities from DXF file.
    drawing_scale: 1 drawing unit = 1/scale meters (default 1:100 → 1 unit = 0.01 m)
    Returns dict with 'blocks', 'pipes', 'metadata'.
    """
    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()

    block_counts  = Counter()     # block_name → count
    pipe_lengths  = defaultdict(float)  # layer_name → total length

    PIPE_LAYER_RE = re.compile(r"PIPE|DRAIN|STACK|VENT|SUPPLY|SOIL", re.I)

    for entity in msp:
        etype = entity.dxftype()

        if etype == "INSERT":
            block_counts[entity.dxf.name] += 1

        elif etype in ("LINE", "LWPOLYLINE"):
            layer = getattr(entity.dxf, "layer", "0")
            if PIPE_LAYER_RE.search(layer):
                pipe_lengths[layer] += line_length(entity)

    # Parse block names
    blocks = []
    for name, qty in sorted(block_counts.items()):
        info = parse_block_name(name)
        info["qty"] = qty
        blocks.append(info)

    # Convert pipe lengths (drawing units → meters)
    unit_to_m = 1.0 / drawing_scale
    pipes = {layer: round(length * unit_to_m, 2)
             for layer, length in pipe_lengths.items()}

    # Metadata
    all_layers = [layer.dxf.name for layer in doc.layers]
    all_block_defs = [b.name for b in doc.blocks if not b.name.startswith("*")]

    return {
        "blocks":      blocks,
        "pipes":       pipes,
        "layers":      all_layers,
        "block_defs":  all_block_defs,
        "source_file": os.path.basename(dxf_path),
        "total_fittings": sum(b["qty"] for b in blocks),
    }


# ─────────────────────────────────────────────────────────────────────────────
# CONSOLE PRINT
# ─────────────────────────────────────────────────────────────────────────────

def print_summary(result: dict):
    print()
    print("=" * 60)
    print(f"  HULIOT QUANTITY TAKEOFF")
    print(f"  Drawing: {result['source_file']}")
    print("=" * 60)

    # Group by product line
    by_product = defaultdict(list)
    for b in result["blocks"]:
        by_product[b["product"]].append(b)

    for product, items in sorted(by_product.items()):
        print(f"\n{product}")
        print("-" * 40)
        for item in sorted(items, key=lambda x: (x["type"], x["size"])):
            print(f"  {item['type']:25s} {item['size']:15s}  x {item['qty']:4d}")

    if result["pipes"]:
        print(f"\nPIPE RUNS (estimated lengths at 1:{int(1/0.01)} scale)")
        print("-" * 40)
        for layer, length_m in sorted(result["pipes"].items()):
            print(f"  {layer:35s}  {length_m:.1f} m")

    print()
    print(f"  TOTAL FITTINGS: {result['total_fittings']}")
    print("=" * 60)
    print()


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

GREEN_DARK  = "1A5C38"
GREEN_MID   = "2D8A56"
GREEN_LIGHT = "C8F0D8"
GREY_LIGHT  = "F5F5F5"


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def export_excel(result: dict, out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quantity Takeoff"

    # --- Title ---
    ws.merge_cells("A1:H1")
    ws["A1"] = "HULIOT PIPES & FITTINGS — QUANTITY TAKEOFF"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=GREEN_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Source Drawing: {result['source_file']}"
    ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color=GREEN_DARK)
    ws["A2"].fill      = PatternFill("solid", fgColor=GREEN_LIGHT)
    ws["A2"].alignment = Alignment(horizontal="center")

    # --- Header row ---
    headers = ["Sr.", "Block Name (CAD)", "Product Line", "Type",
               "Size", "Qty", "Unit", "Remarks"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=GREEN_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
    ws.row_dimensions[4].height = 20

    # --- Data rows ---
    row = 5
    sr = 1

    by_product = defaultdict(list)
    for b in result["blocks"]:
        by_product[b["product"]].append(b)

    for product in sorted(by_product.keys()):
        items = by_product[product]

        # Product group header
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = f"▶  {product}"
        ws[f"A{row}"].font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        ws[f"A{row}"].fill      = PatternFill("solid", fgColor=GREEN_DARK)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

        for item in sorted(items, key=lambda x: (x["type"], x["size"])):
            fill = PatternFill("solid", fgColor=GREY_LIGHT) if sr % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            values = [sr, item["raw_name"], item["product"],
                      item["type"], item["size"], item["qty"], "No.", ""]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font      = Font(name="Calibri", size=9)
                cell.fill      = fill
                cell.border    = _border()
                cell.alignment = Alignment(horizontal="center" if col in (1, 6) else "left",
                                           vertical="center")
            row += 1
            sr += 1

    # --- Pipe section ---
    if result["pipes"]:
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = "▶  PIPE RUNS (estimated lengths)"
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=GREEN_DARK)
        ws.row_dimensions[row].height = 18
        row += 1

        for layer, length_m in sorted(result["pipes"].items()):
            values = [sr, layer, "—", "Pipe Run", "—", length_m, "m", "Estimated"]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font   = Font(name="Calibri", size=9)
                cell.border = _border()
                cell.alignment = Alignment(horizontal="center" if col in (1, 6) else "left",
                                           vertical="center")
            row += 1
            sr += 1

    # --- Total row ---
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "TOTAL FITTINGS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=GREEN_MID)
    ws[f"A{row}"].alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=6, value=result["total_fittings"]).font = Font(bold=True, size=11)

    # --- Column widths ---
    widths = [5, 42, 20, 20, 14, 8, 8, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(out_path)
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# CSV EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_csv(result: dict, out_path: str):
    import csv
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Block Name", "Product Line", "Type", "Size", "Qty"])
        for b in result["blocks"]:
            w.writerow([b["raw_name"], b["product"], b["type"], b["size"], b["qty"]])


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 extract_qty.py <path_to.dxf> [drawing_scale]")
        print("  drawing_scale: denominator only, e.g. 100 for 1:100 (default 100)")
        sys.exit(1)

    dxf_path = sys.argv[1]
    scale    = float(sys.argv[2]) if len(sys.argv) > 2 else 100.0

    if not os.path.exists(dxf_path):
        print(f"ERROR: File not found: {dxf_path}")
        sys.exit(1)

    ext = Path(dxf_path).suffix.lower()
    if ext == ".dwg":
        print()
        print("ERROR: DWG format cannot be read directly.")
        print()
        print("Convert to DXF in AutoCAD:")
        print("  File → Save As → AutoCAD DXF (*.dxf)")
        print()
        print("Or use free ODA File Converter:")
        print("  https://www.opendesign.com/guestfiles/oda_file_converter")
        sys.exit(1)

    if ext != ".dxf":
        print(f"ERROR: Unsupported format '{ext}'. Need .dxf")
        sys.exit(1)

    print(f"Reading: {dxf_path}")
    result = extract(dxf_path, scale)

    print_summary(result)

    stem = Path(dxf_path).stem
    out_dir = Path(dxf_path).parent

    # CSV always
    csv_path = out_dir / f"blocks_raw_{stem}.csv"
    export_csv(result, str(csv_path))
    print(f"CSV saved:   {csv_path}")

    # Excel if openpyxl available
    if HAS_EXCEL:
        xl_path = out_dir / f"quantity_takeoff_{stem}.xlsx"
        export_excel(result, str(xl_path))
        print(f"Excel saved: {xl_path}")
    else:
        print("openpyxl not installed — Excel skipped. Install: pip install openpyxl")

    return result


if __name__ == "__main__":
    main()
